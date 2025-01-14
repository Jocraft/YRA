import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage

# Models
from accounts.models import Student
from .models import TestSession, Program, PathFindingLog
# (If your "Program" model is in "course.models", then import from there:
# from course.models import Program
# Adjust as needed.)

from .questions import QUESTIONS
from .llm_analyzer import analyze_answer  # Assume you have a custom LLM logic

################################
# Logging helper
################################
def add_log_entry(student, message):
    """
    Helper to create a log entry in PathFindingLog.
    """
    PathFindingLog.objects.create(student=student, message=message)


################################
# VIEWS
################################

def path_finding_home(request):
    """
    Display all students, allow selecting one, or do a bulk upload if none selected.
    Shows a log of recent events.
    """
    students = Student.objects.all()

    selected_student_id = request.GET.get('student_id')
    selected_student = None
    existing_session = None
    enrollment_date = None
    latest_test = None

    if selected_student_id:
        selected_student = get_object_or_404(Student, pk=selected_student_id)
        # Grab the most recent session if multiple are allowed
        existing_session = TestSession.objects.filter(student=selected_student).first()
        user = selected_student.student
        enrollment_date = user.enrollment_date
        latest_test = TestSession.objects.filter(student=selected_student).first()

    # If we had a bulk_upload_log stored in session, retrieve it
    bulk_upload_log = request.session.pop('bulk_upload_log', None)

    # Retrieve the 20 most recent logs
    all_logs = PathFindingLog.objects.select_related('student').order_by('-timestamp')[:20]

    context = {
        'students': students,
        'selected_student': selected_student,
        'existing_session': existing_session,
        'student_id': selected_student_id,
        'enrollment_date': enrollment_date,
        'latest_test': latest_test,
        'bulk_upload_log': bulk_upload_log,
        'all_logs': all_logs,
    }
    return render(request, 'path_finding/path_finding_home.html', context)


def create_test_session(request, student_id):
    """
    Creates a new TestSession for the given student (manual creation).
    """
    student = get_object_or_404(Student, pk=student_id)
    session = TestSession.objects.create(student=student)
    
    # Log
    add_log_entry(student, f"Created a new TestSession (id={session.id}) manually for {student}.")

    return redirect('fill_test', session_id=session.id)


def fill_test_view(request, session_id):
    """
    Page where the student manually answers the test (HTML form).
    """
    session = get_object_or_404(TestSession, pk=session_id)

    if request.method == 'POST':
        # Map question numbers to model fields
        field_mapping = {
            1: 'q1_numbers',
            2: 'q2_math',
            3: 'q3_problem_solving',
            4: 'q4_future_job',
            5: 'q5_tools',
            6: 'q6_tech_interest',
            7: 'q7_learning',
            8: 'q8_computer_skills',
            9: 'q9_projects',
            10: 'q10_goals'
        }

        for question in QUESTIONS:
            answer_text = request.POST.get(f'q{question["order"]}')
            if answer_text:
                field_name = field_mapping[question['order']]
                setattr(session, field_name, answer_text)

        session.is_complete = True
        session.save()

        # Log
        add_log_entry(session.student, f"Manually filled answers for TestSession (id={session.id}).")

        return redirect('generate_results', session_id=session.id)

    return render(request, 'path_finding/fill_test.html', {
        'session': session,
        'questions': QUESTIONS
    })


def generate_results_view(request, session_id):
    """
    Analyzes the student's answers using LLM and recommends programs.
    """
    session = get_object_or_404(TestSession, pk=session_id)
    programs = Program.objects.all()

    field_mapping = {
        1: 'q1_numbers',
        2: 'q2_math',
        3: 'q3_problem_solving',
        4: 'q4_future_job',
        5: 'q5_tools',
        6: 'q6_tech_interest',
        7: 'q7_learning',
        8: 'q8_computer_skills',
        9: 'q9_projects',
        10: 'q10_goals'
    }

    # Collect Q&A
    qa_pairs = []
    for i in range(1, 11):
        question = next((q for q in QUESTIONS if q['order'] == i), None)
        if question:
            field_name = field_mapping[i]
            answer = getattr(session, field_name, "")
            qa_pairs.append({
                'question': question['text'],
                'answer': answer,
                'question_id': i
            })

    # Analyze with your LLM (pseudo-code)
    if qa_pairs:
        answer_summary = "\n\n".join([
            f"Question {qa['question_id']}: {qa['question']}\nStudent's Answer: {qa['answer']}"
            for qa in qa_pairs
        ])
        result = analyze_answer(
            question="Based on the student's answers, recommend suitable programs.",
            answer=answer_summary,
            programs=programs
        )
        session.llm_analysis = result
        session.save()
    else:
        result = "No answers found to analyze."

    return render(request, 'path_finding/results.html', {
        'session': session,
        'qa_pairs': qa_pairs,
        'result': result
    })


def upload_answers_view(request, student_id):
    """
    Single upload for one specific student: 
    [Timestamp, Name, ID, Q1..Q10] → only 1 row or more. 
    """
    student = get_object_or_404(Student, pk=student_id)
    existing_session = TestSession.objects.filter(student=student).first()
    
    created_new_session = False
    if not existing_session:
        existing_session = TestSession.objects.create(student=student)
        created_new_session = True

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "No file was uploaded.")
            return redirect('upload_answers', student_id=student.id)

        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        uploaded_file_path = fs.path(filename)

        try:
            wb = openpyxl.load_workbook(uploaded_file_path)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
            return redirect('upload_answers', student_id=student.id)

        row_count = 0
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Skip header
            if idx == 1:
                continue
            row_count += 1

            if len(row) < 13:
                messages.warning(request, f"Row {idx}: only {len(row)} columns found; expected >= 13. Skipped.")
                continue

            # row[2] -> ID in the Excel (not used for logic here, because we already know the student)
            answers = row[3:13]  # 10 answers
            try:
                existing_session.q1_numbers         = answers[0] or ""
                existing_session.q2_math            = answers[1] or ""
                existing_session.q3_problem_solving = answers[2] or ""
                existing_session.q4_future_job      = answers[3] or ""
                existing_session.q5_tools           = answers[4] or ""
                existing_session.q6_tech_interest   = answers[5] or ""
                existing_session.q7_learning        = answers[6] or ""
                existing_session.q8_computer_skills = answers[7] or ""
                existing_session.q9_projects        = answers[8] or ""
                existing_session.q10_goals          = answers[9] or ""

                existing_session.is_complete = True
                existing_session.save()

            except Exception as e:
                messages.warning(request, f"Row {idx}: error saving answers: {str(e)}")

        if row_count > 0:
            messages.success(request, "Answers uploaded successfully!")
        else:
            messages.warning(request, "No valid data rows found in the Excel file.")

        # Log
        if created_new_session:
            add_log_entry(student, f"Single-file upload created new TestSession (id={existing_session.id}).")
        else:
            add_log_entry(student, f"Single-file upload updated TestSession (id={existing_session.id}).")

        return redirect('generate_results', session_id=existing_session.id)

    # If GET request
    return render(request, 'path_finding/upload_answers.html', {
        'student': student,
        'existing_session': existing_session
    })


def upload_bulk_answers_view(request):
    """
    Handles bulk uploading an Excel file with multiple rows:
    [Timestamp, Name, ID, Q1..Q10].
    For each row, we find the student by that ID, create/update a TestSession, and log it.
    """
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "No file was uploaded.")
            return redirect('upload_bulk_answers')

        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        uploaded_file_path = fs.path(filename)

        try:
            wb = openpyxl.load_workbook(uploaded_file_path)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
            return redirect('upload_bulk_answers')

        row_count = 0
        updated_count = 0
        bulk_upload_log = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Skip header
            if idx == 1:
                continue
            row_count += 1

            if len(row) < 13:
                messages.warning(request, f"Row {idx}: only {len(row)} columns found; expected >= 13. Skipped.")
                continue

            excel_id = row[2]  # Student ID
            answers = row[3:13]  # Q1..Q10

            if not excel_id:
                messages.warning(request, f"Row {idx}: missing ID. Skipped.")
                continue

            # Find the student
            try:
                student = Student.objects.get(pk=excel_id)
            except Student.DoesNotExist:
                messages.warning(request, f"Row {idx}: no Student found with ID={excel_id}. Skipped.")
                continue

            # Either update existing or create new
            session = TestSession.objects.filter(student=student).first()
            created_new_session = False
            if not session:
                session = TestSession.objects.create(student=student)
                created_new_session = True

            # Fill answers
            try:
                session.q1_numbers         = answers[0] or ""
                session.q2_math            = answers[1] or ""
                session.q3_problem_solving = answers[2] or ""
                session.q4_future_job      = answers[3] or ""
                session.q5_tools           = answers[4] or ""
                session.q6_tech_interest   = answers[5] or ""
                session.q7_learning        = answers[6] or ""
                session.q8_computer_skills = answers[7] or ""
                session.q9_projects        = answers[8] or ""
                session.q10_goals          = answers[9] or ""

                session.is_complete = True
                session.save()
                updated_count += 1

                if created_new_session:
                    msg = f"Bulk upload: Created new session (id={session.id}) for {student}."
                else:
                    msg = f"Bulk upload: Updated session (id={session.id}) for {student}."
                bulk_upload_log.append(msg)

                # Also log in DB
                add_log_entry(student, msg)

            except Exception as e:
                messages.warning(request, f"Row {idx}: error saving answers: {str(e)}")

        if row_count > 0:
            messages.success(request, f"Processed {row_count} row(s). Updated/created {updated_count} session(s).")
        else:
            messages.warning(request, "No valid data rows found in the Excel file.")

        # Store a textual log in the session so we can show it in home template
        request.session['bulk_upload_log'] = bulk_upload_log

        return redirect('path_finding_home')

    # GET: show form
    return render(request, 'path_finding/upload_bulk_answers.html')


def delete_logs(request):
    """
    Deletes all entries in PathFindingLog.
    """
    PathFindingLog.objects.all().delete()
    messages.success(request, "All logs have been deleted.")
    return redirect('path_finding_home')